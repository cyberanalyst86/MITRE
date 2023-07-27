import PySimpleGUI as sg
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def color_cell(cell, mitre_id_list):
    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    if cell.value != None:

        cell.border = border

    for id in mitre_id_list:

        try:
            if id in cell.value:
                cell.fill =  PatternFill("solid", fgColor="00FFFF00")


        except TypeError:

            continue


def get_metadata(workbook):
    """Prints the number of rows and columns of an Excel workbook.

    Args:
        workbook (str): The name of the Excel workbook.
    """

    sheet = workbook.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    return row_count, column_count


if __name__ == "__main__":

    #-------------------------------Create GUI-------------------------------

    sg.theme("DarkTeal2")
    layout = [[sg.T("")], [sg.Text("Mitre Technique ID File: "), sg.Input(), sg.FileBrowse(key="-IN-")], [sg.Text('Output File Name:', size =(16, 1), ), sg.InputText(visible=True, enable_events=True, key='fig_path'),
        sg.FileSaveAs(
            file_types=('XLSX','.xlsx'),  # TODO: better names
        )] , [sg.Button("Submit for processing")]]
    ###Building Window
    window = sg.Window('Mitre Att&ck Matrix (Enterprise) Visualiser', layout, size=(600, 150))

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        elif event == "Submit for processing":
            dialog_input = values

            break

    #-------------------------------Get Parameters from GUI-------------------------------

    input_filepath = dialog_input['-IN-']
    output_filename = dialog_input['fig_path']

    mitre_template_filepath = '//'.join(str(e) for e in input_filepath.split("/")[:-1])

    mitre_template_file = mitre_template_filepath + "//" + "mitre_matrix_template.xlsx"


    #-------------------------------Read Mitre Technique ID------------------------------

    df = pd.read_excel(input_filepath, header=None)

    mitre_id_list = df.loc[:, 0].values.tolist()

    #-------------------------------Read Mitre Template File-------------------------------

    excel_column_list = ['A','B', 'C', 'D', 'E' , 'F' , 'G', 'H' , 'I' , 'J' , 'K', 'L', 'M', 'N']


    wb = openpyxl.load_workbook(mitre_template_file)

    row_count, column_count = get_metadata(wb)

    ws = wb['Sheet1']

    for c in range(len(excel_column_list)):

        for r in range(row_count):
            row = r + 2
            column = c + 1

            # print("(" + str(row) + ", " + str(column) + ")")

            # set the height of the row
            ws.row_dimensions[row].height = 50

            # set the width of the column
            ws.column_dimensions[excel_column_list[c]].width = 15

            font = openpyxl.styles.Font(size=10, bold=False)
            alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center', wrapText=True)

            ws.cell(row=row - 1, column=column).font = font
            ws.cell(row=row - 1, column=column).alignment = alignment

            color_cell(ws.cell(row=row, column=column), mitre_id_list)

    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    # Set the border for the sheet
    ws.border = border

    wb.save(output_filename )
