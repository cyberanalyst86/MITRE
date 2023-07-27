import pandas as pd
import openpyxl
import re
from openpyxl.styles import PatternFill


# Function to highlight cells based on Mitre Technique ID
def color_cell(cell, mitre_id_list):
    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    if cell.value != None:
        cell.border = border

    for id in mitre_id_list:

        try:
            if id in cell.value:
                cell.fill = PatternFill("solid", fgColor="00FFFF00")


        except TypeError:

            continue


# Function to get the number of rows and columns from Mitre ATT&CK Enterprise Matrix Template
def get_metadata(workbook):
    """Prints the number of rows and columns of an Excel workbook.

    Args:
        workbook (str): The name of the Excel workbook.
    """

    sheet = workbook.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    return row_count, column_count


# Main Function
def main():

    excel_column_list = ['A','B', 'C', 'D', 'E' , 'F' , 'G', 'H' , 'I' , 'J' , 'K', 'L', 'M', 'N']

    wb = openpyxl.load_workbook("C:\\Users\\ongye\\Downloads\\test_python_code\\mitre_matrix_template.xlsx")

    row_count, column_count = get_metadata(wb)

    ws = wb['Sheet1']

    name = input("Enter APT or Malware name (this will be used for output filename) : ")
    print("\n")

    mitre_id_string = input("Enter Mitre Technique IDs delimited by comma: ")
    print("\n")

    mitre_id_list = mitre_id_string.replace("'", "").replace(" ", "").split(",")

    # mitre_id_list = ['T1548', 'T1134', 'T1087', 'T1059', 'T1486', 'T1491', 'T1561', 'T1083', 'T1222', 'T1070', 'T1490', 'T1570', 'T1112', 'T1135', 'T1069', 'T1018', 'T1489', 'T1082', 'T1033', 'T1047']

    print("Mitre ID list = ", str(mitre_id_list))
    print("\n")

    for c in range(len(excel_column_list)):

        for r in range(row_count):
            row = r + 2
            column = c + 1

            #print("(" + str(row) + ", " + str(column) + ")")

            # set the height of the row
            ws.row_dimensions[row].height = 50

            # set the width of the column
            ws.column_dimensions[excel_column_list[c]].width = 15

            font = openpyxl.styles.Font(size=10, bold=False)
            alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center', wrapText=True)

            ws.cell(row=row-1, column=column).font = font
            ws.cell(row=row-1, column=column).alignment = alignment

            color_cell(ws.cell(row=row, column=column), mitre_id_list)

    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    # Set the border for the sheet
    ws.border = border

    outputfilename = str(name) + "_mitre_matrix_enterprise.xlsx"

    wb.save(outputfilename)

    print("Mitre matrix generated successfully")


if __name__ == "__main__":
    main()