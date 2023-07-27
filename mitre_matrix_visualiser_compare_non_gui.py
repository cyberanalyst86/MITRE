import pandas as pd
import openpyxl
import re
from openpyxl.styles import PatternFill
#https://openpyxl.readthedocs.io/en/stable/styles.html

#Function to highlight cells based on Mitre Technique ID
def color_cell(cell, malware1_mitre_id_list, malware2_mitre_id_list):
    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    if cell.value != None:

        cell.border = border

    id_list = []
    id2_list = []

    for id in malware1_mitre_id_list:

        for id2 in malware2_mitre_id_list:

            try:
                if id in cell.value and id2 in cell.value:
                    #orange
                    cell.fill =  PatternFill("solid", fgColor="00FF6600")

                    id_list.append(id)
                    id2_list.append(id2)

            except TypeError:

                continue

    for id in id_list:

        malware1_mitre_id_list.remove(id)

    for id2 in id2_list:
        malware2_mitre_id_list.remove(id2)

    for id in malware1_mitre_id_list:

        try:
            if id in cell.value:
                # yellow
                cell.fill = PatternFill("solid", fgColor="00FFFF00")

        except TypeError:

            continue

    for id2 in malware2_mitre_id_list:

        try:
            if id2 in cell.value:
                # green
                cell.fill = PatternFill("solid", fgColor="0000FF00")

        except TypeError:

            continue

#Function to get the number of rows and columns from Mitre ATT&CK Enterprise Matrix Template
def get_metadata(workbook):
    """Prints the number of rows and columns of an Excel workbook.

    Args:
        workbook (str): The name of the Excel workbook.
    """

    sheet = workbook.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    return row_count, column_count

#Main Function
def main():

    excel_column_list = ['A','B', 'C', 'D', 'E' , 'F' , 'G', 'H' , 'I' , 'J' , 'K', 'L', 'M', 'N']


    wb = openpyxl.load_workbook("C:\\Users\\ongye\\Downloads\\test_python_code\\mitre_matrix_template.xlsx")

    row_count, column_count = get_metadata(wb)

    ws = wb['Sheet1']

    name1 = input("Enter 1st APT or Malware name (this will be used for output filename) : ")
    print("\n")
    mitre_id_string1 = input("Enter 1st set of Mitre Technique IDs delimited by comma: ")
    print("\n")
    malware1_mitre_id_list = mitre_id_string1.replace("'", "").replace(" ", "").split(",")

    #malware1_mitre_id_list = ['T1564', 'T1055.003', 'T1082', 'T1059', 'T1547.001', 'T1560', 'T1560.001', 'T1543.003',
     #                    'T1134.001', 'T1547', 'T1569', 'T1497.001', 'T1016', 'T1587', 'T1134', 'T1012', 'T1055',
     #                    'T1071.001', 'T1027.002', 'T1548', 'T1021.002', 'T1569.002', 'T1562', 'T1490', 'T1070',
     #                    'T1587.002', 'T1497', 'T1140', 'T1518', 'T1543', 'T1070.004', 'T1213', 'T1548.002', 'T1489',
     #                    'T1071', 'T1564.003', 'T1485', 'T1057', 'T1486', 'T1059.001', 'T1056', 'T1021', 'T1055.012',
     #                    'T1027', 'T1033', 'T1056.001', 'T1010', 'T1047', 'T1083', 'T1113', 'T1007', 'T1529', 'T1135',
     #                    'T1087', 'T1112', 'T1562.002']

    name2 = input("Enter 2nd APT or Malware name (this will be used for output filename) : ")
    print("\n")

    mitre_id_string2 = input("Enter 2nd set of Mitre Technique IDs delimited by comma: ")
    print("\n")

    malware2_mitre_id_list = mitre_id_string2.replace("'", "").replace(" ", "").split(",")

    #malware2_mitre_id_list =['T1135', 'T1090', 'T1087', 'T1057', 'T1082', 'T1007', 'T1027', 'T1562', 'T1070.004', 'T1033', 'T1090.003', 'T1543.003', 'T1046', 'T1569.002', 'T1071', 'T1112', 'T1548.002', 'T1489', 'T1070', 'T1095', 'T1543', 'T1012', 'T1059', 'T1529', 'T1070.006', 'T1083', 'T1562.002', 'T1071.001', 'T1134', 'T1564', 'T1140', 'T1047', 'T1134.001', 'T1569', 'T1055', 'T1548', 'T1490', 'T1486', 'T1497.001', 'T1497']

    print(str(name1) + " mitre id : " + str(malware1_mitre_id_list))
    print(str(name2) + " mitre id : " + str(malware2_mitre_id_list) + "\n")

    print("************!!!Note!!!************\n")
    print("Common Mitre Techniques highlighted in 'Orange'")
    print("1st set of unique Mitre Techniques highlighted in 'Yellow'")
    print("2nd set of unique Mitre Techniques highlighted in 'Green'\n")

    for c in range(len(excel_column_list)):

        for r in range(row_count):

            row = r + 2
            column = c + 1

            #print("(" + str(row) + ", " + str(column) + ")")

            # set the height of the row
            ws.row_dimensions[row].height = 50

            # set the width of the column
            ws.column_dimensions[excel_column_list[c]].width = 15

            font = openpyxl.styles.Font(size=10, bold=False)
            alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center', wrapText=True)

            ws.cell(row=row - 1, column=column).font = font
            ws.cell(row=row - 1, column=column).alignment = alignment

            color_cell(ws.cell(row=row, column=column), malware1_mitre_id_list, malware2_mitre_id_list)

    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    # Set the border for the sheet
    ws.border = border

    outputfilename = str(name1) + "_" + str(name2) + "_mitre_matrix_enterprise.xlsx"

    wb.save(outputfilename)

    print("Mitre matrix generated successfully as " + str(outputfilename))

if __name__=="__main__":
    main()
