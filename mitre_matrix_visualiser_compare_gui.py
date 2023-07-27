import PySimpleGUI as sg
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

#https://openpyxl.readthedocs.io/en/stable/styles.html
def color_cell(cell, malware1_mitre_id_list, malware2_mitre_id_list):
    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    if cell.value != None:

        cell.border = border

    id_list = []
    id2_list = []

    for id in malware1_mitre_id_list:

        for id2 in malware2_mitre_id_list:

            try:
                if id in cell.value and id2 in cell.value:
                    #orange
                    cell.fill =  PatternFill("solid", fgColor="00FF6600") #orange

                    id_list.append(id)
                    id2_list.append(id2)

            except TypeError:

                continue

    for id in id_list:

        malware1_mitre_id_list.remove(id)

    for id2 in id2_list:
        malware2_mitre_id_list.remove(id2)

    for id in malware1_mitre_id_list:

        try:
            if id in cell.value:
                # yellow
                cell.fill = PatternFill("solid", fgColor="00FFFF00") #yellow

        except TypeError:

            continue

    for id2 in malware2_mitre_id_list:

        try:
            if id2 in cell.value:
                # green
                cell.fill = PatternFill("solid", fgColor="0000FF00") #green

        except TypeError:

            continue


def get_metadata(workbook):
    """Prints the number of rows and columns of an Excel workbook.

    Args:
        workbook (str): The name of the Excel workbook.
    """

    sheet = workbook.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    return row_count, column_count


if __name__ == "__main__":

    #-------------------------------Create GUI-------------------------------

    sg.theme("DarkTeal2")
    layout = [[sg.Text('ID Set #1 - will be highlighted yellow', enable_events=True,
   key='-TEXT-', font=('Arial Bold', 10),
   expand_x=True, justification='left')], \
              [sg.Text('ID Set #2 - will be highlighted green ', enable_events=True,
                        key='-TEXT-', font=('Arial Bold', 10),
                        expand_x=True, justification='left')], \
                  [sg.Text('ID Set #1 and Set #2 overlapped - will be highligted orange ', enable_events=True,
                            key='-TEXT-', font=('Arial Bold', 10),
                            expand_x=True, justification='left')],
                [sg.T("")], [sg.Text("Mitre Technique ID File Set #1: "), sg.Input(), sg.FileBrowse(key="-IN1-")], \
              [sg.T("")], [sg.Text("Mitre Technique ID File Set #2: "), sg.Input(), sg.FileBrowse(key="-IN2-")],
              [sg.Text('Output File Name:', size =(16, 1), ), sg.InputText(visible=True, enable_events=True, key='fig_path'),
        sg.FileSaveAs(
            file_types=('XLSX','.xlsx'),  # TODO: better names
        )] , [sg.Button("Submit for processing")]]
    ###Building Window
    window = sg.Window('Mitre Att&ck Matrix (Enterprise) Visualiser (Comparison)', layout, size=(600, 300))

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        elif event == "Submit for processing":
            dialog_input = values

            break

    #-------------------------------Get Parameters from GUI-------------------------------

    input_filepath1 = dialog_input['-IN1-']
    input_filepath2 = dialog_input['-IN2-']
    output_filename = dialog_input['fig_path']

    mitre_template_filepath = '//'.join(str(e) for e in input_filepath1.split("/")[:-1])

    mitre_template_file = mitre_template_filepath + "//" + "mitre_matrix_template.xlsx"

    #-------------------------------Read Mitre Technique ID------------------------------

    df1 = pd.read_excel(input_filepath1, header=None)
    df2 = pd.read_excel(input_filepath2, header=None)

    malware1_mitre_id_list = df1.loc[:, 0].values.tolist()
    malware2_mitre_id_list = df2.loc[:, 0].values.tolist()


    #-------------------------------Read Mitre Template File-------------------------------

    excel_column_list = ['A','B', 'C', 'D', 'E' , 'F' , 'G', 'H' , 'I' , 'J' , 'K', 'L', 'M', 'N']


    wb = openpyxl.load_workbook(mitre_template_file)

    row_count, column_count = get_metadata(wb)

    ws = wb['Sheet1']

    for c in range(len(excel_column_list)):

        for r in range(row_count):
            row = r + 2
            column = c + 1

            # print("(" + str(row) + ", " + str(column) + ")")

            # set the height of the row
            ws.row_dimensions[row].height = 50

            # set the width of the column
            ws.column_dimensions[excel_column_list[c]].width = 15

            font = openpyxl.styles.Font(size=10, bold=False)
            alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center', wrapText=True)

            ws.cell(row=row - 1, column=column).font = font
            ws.cell(row=row - 1, column=column).alignment = alignment

            color_cell(ws.cell(row=row, column=column), malware1_mitre_id_list, malware2_mitre_id_list)

    # Create a border object
    border = openpyxl.styles.Border()
    border.top = openpyxl.styles.Side(style='thin')
    border.bottom = openpyxl.styles.Side(style='thin')
    border.left = openpyxl.styles.Side(style='thin')
    border.right = openpyxl.styles.Side(style='thin')

    # Set the border for the sheet
    ws.border = border

    wb.save(output_filename)
